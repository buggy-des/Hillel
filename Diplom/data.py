import openpyxl
from person import Person


class PersonData:
    def __init__(self):
        self.people = []

    def add_person(self, person: Person):
        self.people.append(person)

    def load_from_file(self, filename: str):
        try:
            wb = openpyxl.load_workbook(filename)
            ws = wb.active
            self.people = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                first_name, patronymic, surname, gender, _, date_of_birth_str, date_of_death_str = row
                date_of_death = None if date_of_death_str == "-" else date_of_death_str
                person = Person(first_name, gender, date_of_birth_str, date_of_death, surname, patronymic)
                self.people.append(person)
            print(f"Файл {filename} завантажено.")
        except Exception:
            print(f"Помилка завантаження файлу, можливо не додано розширення (.xlsx,.xlsm,.xltx,.xltm)")

    def find_person(self, request: str):
        request = request.lower().strip()
        found_people = [person for person in self.people if request in person.first_name.lower() or
                        request in (person.surname.lower() if person.surname else '') or
                        request in (person.patronymic.lower() if person.patronymic else '')]
        if not found_people:
            print(f"Людина {request} відсутня у базі")
        return [str(person) for person in found_people]

    def save_to_file(self, filename: str):
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(['ім\'я', 'По-батькові', 'Прізвище', 'Стать', 'Вік', 'Дата народження', 'Дата смерті'])
            for person in self.people:
                person_data = person.person_info()
                person_data = ['-' if value is None else value for value in person_data]
                ws.append(person_data)
            wb.save(filename)
            print(f"Файл {filename} збережено.")
        except Exception as e:
            print(f"Помилка збереження файлу {e}, можливо не додано розширення (.xlsx,.xlsm,.xltx,.xltm)")

    def __str__(self):
        return '\n' + ('-' * 100 + '\n').join(str(person) for person in self.people)
