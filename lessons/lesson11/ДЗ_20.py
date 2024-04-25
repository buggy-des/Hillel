import csv

import openpyxl

with open('example.csv', 'r', newline='', encoding='UTF-8') as file:
    file_reader = csv.reader(file, delimiter=',')
    data = list(file_reader)

transposed_data = list(map(list, zip(*data)))

for row in data:
    del row[2]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Data'

headers = transposed_data[0]
for col_num, header in enumerate(headers, 1):
    ws.cell(row=1, column=col_num, value=header)

for row_num, row_data in enumerate(transposed_data[1:], 2):
    for col_num, cell_data in enumerate(row_data, 1):
        ws.cell(row=row_num, column=col_num, value=cell_data)

wb.save('example.xlsx')
